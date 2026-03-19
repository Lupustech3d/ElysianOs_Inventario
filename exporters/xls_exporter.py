import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ── Cores ElysianOS ────────────────────────────────────
COLOR_BG_DARK  = "1A1A2E"
COLOR_BG_MID   = "16213E"
COLOR_BG_LIGHT = "0F3460"
COLOR_TEAL     = "00BCD4"
COLOR_WHITE    = "FFFFFF"
COLOR_GRAY     = "7F8C8D"
COLOR_GREEN    = "00BC64"
COLOR_RED      = "E74C3C"
COLOR_YELLOW   = "F39C12"
COLOR_ROW_EVEN = "141D35"
COLOR_ROW_ODD  = "16213E"


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def make_font(bold=False, color=COLOR_WHITE, size=10):
    return Font(bold=bold, color=color, size=size, name="Segoe UI")


def make_border():
    side = Side(style="thin", color="1E3A5F")
    return Border(left=side, right=side, top=side, bottom=side)


def make_center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def make_left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols, bg=COLOR_BG_LIGHT,
                     font_color=COLOR_TEAL, height=22):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = make_fill(bg)
        cell.font = make_font(bold=True, color=font_color, size=9)
        cell.alignment = make_center()
        cell.border = make_border()
    ws.row_dimensions[row].height = height


def style_data_row(ws, row, cols, even=True):
    bg = COLOR_ROW_EVEN if even else COLOR_ROW_ODD
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = make_fill(bg)
        cell.font = make_font(size=9)
        cell.alignment = make_left()
        cell.border = make_border()
    ws.row_dimensions[row].height = 18


def add_section_title(ws, row, title, cols, col_start=1):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_start + cols - 1
    )
    cell = ws.cell(row=row, column=col_start)
    cell.value = title
    cell.fill = make_fill(COLOR_BG_DARK)
    cell.font = make_font(bold=True, color=COLOR_TEAL, size=10)
    cell.alignment = make_left()
    ws.row_dimensions[row].height = 24
    return row + 1


def add_logo(ws, logo_path, row=1, col=1):
    try:
        img = XLImage(logo_path)
        img.width  = 55
        img.height = 55
        ws.add_image(img, f"{get_column_letter(col)}{row}")
        return True
    except:
        return False


def get_logo_path():
    base = getattr(sys, '_MEIPASS',
                   os.path.dirname(os.path.abspath(__file__)))
    return os.path.normpath(os.path.join(base, "..", "assets", "logo.png"))


# ══════════════════════════════════════════════════════
#  CAPA
# ══════════════════════════════════════════════════════

def build_capa(wb, data):
    ws = wb.active
    ws.title = "Capa"
    ws.sheet_view.showGridLines = False

    # Larguras
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 4
    ws.column_dimensions["F"].width = 4

    # Fundo escuro
    for r in range(1, 50):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = make_fill(COLOR_BG_DARK)

    # Logo na coluna C linha 2
    logo_path = get_logo_path()
    add_logo(ws, logo_path, row=2, col=3)
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 22

    # Título
    ws.merge_cells("C5:F5")
    c = ws["C5"]
    c.value = "ElysianOS"
    c.font = Font(bold=True, color=COLOR_TEAL, size=22, name="Segoe UI")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = make_fill(COLOR_BG_DARK)
    ws.row_dimensions[5].height = 35

    ws.merge_cells("C6:F6")
    c = ws["C6"]
    c.value = "Sistema de Inventário Integrado"
    c.font = Font(color=COLOR_GRAY, size=12, name="Segoe UI")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = make_fill(COLOR_BG_DARK)
    ws.row_dimensions[6].height = 22

    # Separador teal
    ws.merge_cells("C8:F8")
    ws["C8"].fill = make_fill(COLOR_TEAL)
    ws.row_dimensions[8].height = 3

    # Dados
    system   = data.get("system", {})
    network  = data.get("network", {})
    licenses = data.get("licenses", {})

    items = [
        ("RELATÓRIO DE INVENTÁRIO", None),
        (None, None),
        ("Colaborador",         system.get("collaborator", "—")),
        ("Usuário Windows",     system.get("windows_user", "—")),
        ("Máquina",             system.get("machine_name", "—")),
        ("Fabricante / Modelo", f"{system.get('manufacturer','—')} "
                                f"{system.get('model','—')}"),
        ("Nº de Série",         system.get("serial_number", "—")),
        ("Sistema Operacional", licenses.get("windows_edition", "—")),
        ("IP Principal",        network.get("main_ip", "—")),
        ("Data / Hora",         system.get("inventory_datetime", "—")),
        (None, None),
        ("Desenvolvido por",    "Luis Mesquitela"),
    ]

    row = 10
    for label, value in items:
        if label == "RELATÓRIO DE INVENTÁRIO":
            ws.merge_cells(f"C{row}:F{row}")
            c = ws.cell(row=row, column=3)
            c.value = label
            c.font = Font(bold=True, color=COLOR_WHITE,
                          size=13, name="Segoe UI")
            c.alignment = Alignment(horizontal="center",
                                    vertical="center")
            c.fill = make_fill(COLOR_BG_LIGHT)
            ws.row_dimensions[row].height = 28
        elif label is None:
            ws.row_dimensions[row].height = 8
        else:
            # Label — coluna C
            cl = ws.cell(row=row, column=3, value=label)
            cl.font  = Font(bold=True, color=COLOR_TEAL,
                            size=10, name="Segoe UI")
            cl.fill  = make_fill(COLOR_BG_LIGHT)
            cl.alignment = Alignment(horizontal="left",
                                     vertical="center",
                                     wrap_text=False)
            cl.border = make_border()

            # Valor — coluna D mesclada até F
            ws.merge_cells(f"D{row}:F{row}")
            cv = ws.cell(row=row, column=4, value=value)
            cv.font  = Font(color=COLOR_WHITE, size=10,
                            name="Segoe UI")
            cv.fill  = make_fill(COLOR_BG_MID)
            cv.alignment = Alignment(horizontal="left",
                                     vertical="center",
                                     wrap_text=False)
            cv.border = make_border()
            ws.row_dimensions[row].height = 20
        row += 1

    # Footer
    ws.merge_cells(f"C{row+2}:F{row+2}")
    c = ws.cell(row=row+2, column=3)
    c.value = (f"ElysianOS {datetime.now().year} "
               f"| Desenvolvido por Luis Mesquitela")
    c.font = Font(color=COLOR_GRAY, size=8,
                  name="Segoe UI", italic=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = make_fill(COLOR_BG_DARK)


# ══════════════════════════════════════════════════════
#  SISTEMA
# ══════════════════════════════════════════════════════

def build_sistema(wb, data):
    ws = wb.create_sheet("Sistema")
    ws.sheet_view.showGridLines = False
    for r in range(1, 50):
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = make_fill(COLOR_BG_DARK)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45

    system = data.get("system", {})
    logo_path = get_logo_path()
    add_logo(ws, logo_path, row=1, col=4)

    row = add_section_title(ws, 1, "💻  INFORMAÇÕES DO SISTEMA", 2)

    fields = [
        ("Nome da Máquina",     system.get("machine_name", "—")),
        ("Colaborador",         system.get("collaborator", "—")),
        ("Usuário Windows",     system.get("windows_user", "—")),
        ("Sistema Operacional", system.get("os_full", "—")),
        ("Arquitetura",         system.get("architecture", "—")),
        ("Tipo de Equipamento", system.get("pc_type", "—")),
        ("Fabricante",          system.get("manufacturer", "—")),
        ("Modelo",              system.get("model", "—")),
        ("Número de Série",     system.get("serial_number", "—")),
        ("Data do Inventário",  system.get("inventory_datetime", "—")),
    ]

    style_header_row(ws, row, 2)
    ws.cell(row=row, column=1).value = "Campo"
    ws.cell(row=row, column=2).value = "Valor"
    row += 1

    for i, (label, value) in enumerate(fields):
        style_data_row(ws, row, 2, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = make_font(bold=True,
                                                     color=COLOR_TEAL)
        ws.cell(row=row, column=2).value = value
        row += 1


# ══════════════════════════════════════════════════════
#  HARDWARE
# ══════════════════════════════════════════════════════

def build_hardware(wb, data):
    ws = wb.create_sheet("Hardware")
    ws.sheet_view.showGridLines = False
    for r in range(1, 80):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = make_fill(COLOR_BG_DARK)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    hardware  = data.get("hardware", {})
    logo_path = get_logo_path()
    add_logo(ws, logo_path, row=1, col=6)

    row = 1

    # CPU
    row = add_section_title(ws, row, "⚙️  PROCESSADOR", 5)
    style_header_row(ws, row, 2)
    ws.cell(row=row, column=1).value = "Campo"
    ws.cell(row=row, column=2).value = "Valor"
    row += 1
    for i, (label, value) in enumerate([
        ("Modelo",          hardware.get("cpu_name", "—")),
        ("Núcleos Físicos", str(hardware.get("cpu_cores_physical", "—"))),
        ("Núcleos Lógicos", str(hardware.get("cpu_cores_logical", "—"))),
        ("Velocidade Máx.", hardware.get("cpu_speed", "—")),
        ("Uso Atual",       hardware.get("cpu_usage", "—")),
    ]):
        style_data_row(ws, row, 2, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = make_font(bold=True,
                                                     color=COLOR_TEAL)
        ws.cell(row=row, column=2).value = value
        row += 1

    row += 1

    # RAM
    row = add_section_title(ws, row, "🧠  MEMÓRIA RAM", 5)
    style_header_row(ws, row, 2)
    ws.cell(row=row, column=1).value = "Campo"
    ws.cell(row=row, column=2).value = "Valor"
    row += 1
    for i, (label, value) in enumerate([
        ("Total",        hardware.get("ram_total", "—")),
        ("Em Uso",       hardware.get("ram_used", "—")),
        ("Disponível",   hardware.get("ram_available", "—")),
        ("Uso %",        hardware.get("ram_percent", "—")),
        ("Slots Usados", str(hardware.get("ram_slots_used", "—"))),
    ]):
        style_data_row(ws, row, 2, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = make_font(bold=True,
                                                     color=COLOR_TEAL)
        ws.cell(row=row, column=2).value = value
        row += 1

    if hardware.get("ram_slots"):
        row += 1
        row = add_section_title(ws, row, "  Slots de RAM", 3)
        style_header_row(ws, row, 3)
        for col, h in enumerate(["Slot","Capacidade","Velocidade"], 1):
            ws.cell(row=row, column=col).value = h
        row += 1
        for i, slot in enumerate(hardware.get("ram_slots", [])):
            style_data_row(ws, row, 3, even=(i % 2 == 0))
            ws.cell(row=row, column=1).value = slot.get("locator","—")
            ws.cell(row=row, column=2).value = slot.get("capacity","—")
            ws.cell(row=row, column=3).value = slot.get("speed","—")
            row += 1

    row += 1

    # Discos
    row = add_section_title(ws, row, "💾  DISCOS FÍSICOS", 5)
    style_header_row(ws, row, 2)
    ws.cell(row=row, column=1).value = "Modelo"
    ws.cell(row=row, column=2).value = "Tamanho"
    row += 1
    for i, disk in enumerate(hardware.get("disks", [])):
        style_data_row(ws, row, 2, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = disk.get("model","—")
        ws.cell(row=row, column=2).value = disk.get("size","—")
        row += 1

    row += 1

    # Partições
    row = add_section_title(ws, row, "📊  PARTIÇÕES", 5)
    style_header_row(ws, row, 5)
    for col, h in enumerate(
            ["Drive","Sistema","Total","Usado","Livre"], 1):
        ws.cell(row=row, column=col).value = h
    row += 1
    for i, part in enumerate(hardware.get("partitions", [])):
        style_data_row(ws, row, 5, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = part.get("drive","—")
        ws.cell(row=row, column=2).value = part.get("filesystem","—")
        ws.cell(row=row, column=3).value = part.get("total","—")
        ws.cell(row=row, column=4).value = part.get("used","—")
        ws.cell(row=row, column=5).value = part.get("free","—")
        row += 1

    row += 1

    # GPU
    row = add_section_title(ws, row, "🎮  GPU", 5)
    style_header_row(ws, row, 2)
    ws.cell(row=row, column=1).value = "Nome"
    ws.cell(row=row, column=2).value = "VRAM"
    row += 1
    for i, gpu in enumerate(hardware.get("gpus", [])):
        style_data_row(ws, row, 2, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = gpu.get("name","—")
        ws.cell(row=row, column=2).value = gpu.get("ram","—")
        row += 1

    row += 1

    # Placa mãe / BIOS
    row = add_section_title(ws, row, "🖥️  PLACA MÃE & BIOS", 5)
    style_header_row(ws, row, 2)
    ws.cell(row=row, column=1).value = "Campo"
    ws.cell(row=row, column=2).value = "Valor"
    row += 1
    for i, (label, value) in enumerate([
        ("Placa Mãe", hardware.get("motherboard","—")),
        ("BIOS",      hardware.get("bios","—")),
    ]):
        style_data_row(ws, row, 2, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = make_font(bold=True,
                                                     color=COLOR_TEAL)
        ws.cell(row=row, column=2).value = value
        row += 1


# ══════════════════════════════════════════════════════
#  REDE
# ══════════════════════════════════════════════════════

def build_rede(wb, data):
    ws = wb.create_sheet("Rede")
    ws.sheet_view.showGridLines = False
    for r in range(1, 60):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = make_fill(COLOR_BG_DARK)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 12

    network   = data.get("network", {})
    logo_path = get_logo_path()
    add_logo(ws, logo_path, row=1, col=6)

    row = 1
    row = add_section_title(ws, row, "🌐  REDE GERAL", 6)
    style_header_row(ws, row, 2)
    ws.cell(row=row, column=1).value = "Campo"
    ws.cell(row=row, column=2).value = "Valor"
    row += 1
    for i, (label, value) in enumerate([
        ("Hostname",      network.get("hostname","—")),
        ("IP Principal",  network.get("main_ip","—")),
        ("MAC Principal", network.get("main_mac","—")),
        ("Gateway",       network.get("main_gateway","—")),
        ("Internet",      network.get("internet","—")),
        ("WiFi SSID",     network.get("wifi_ssid","—")),
    ]):
        style_data_row(ws, row, 2, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = make_font(bold=True,
                                                     color=COLOR_TEAL)
        ws.cell(row=row, column=2).value = value
        row += 1

    row += 1
    row = add_section_title(ws, row, "📡  ADAPTADORES DE REDE", 6)
    style_header_row(ws, row, 6)
    for col, h in enumerate(
            ["Adaptador","IPv4","MAC","Gateway","DNS","DHCP"], 1):
        ws.cell(row=row, column=col).value = h
    row += 1
    for i, a in enumerate(network.get("adapters", [])):
        style_data_row(ws, row, 6, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = a.get("description","—")
        ws.cell(row=row, column=2).value = a.get("ipv4","—")
        ws.cell(row=row, column=3).value = a.get("mac","—")
        ws.cell(row=row, column=4).value = a.get("gateway","—")
        ws.cell(row=row, column=5).value = a.get("dns","—")
        ws.cell(row=row, column=6).value = a.get("dhcp","—")
        row += 1


# ══════════════════════════════════════════════════════
#  SOFTWARE
# ══════════════════════════════════════════════════════

def build_software(wb, data):
    ws = wb.create_sheet("Software")
    ws.sheet_view.showGridLines = False
    for r in range(1, 5):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = make_fill(COLOR_BG_DARK)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12

    software  = data.get("software", {})
    logo_path = get_logo_path()
    add_logo(ws, logo_path, row=1, col=6)

    row = 1
    row = add_section_title(
        ws, row,
        f"📦  SOFTWARES INSTALADOS  —  "
        f"{software.get('software_count',0)} programas",
        5
    )
    style_header_row(ws, row, 5)
    for col, h in enumerate(
            ["Nome","Versão","Fabricante","Instalado em","Tamanho"], 1):
        ws.cell(row=row, column=col).value = h
    row += 1

    for i, sw in enumerate(software.get("software_list", [])):
        bg = COLOR_ROW_EVEN if i % 2 == 0 else COLOR_ROW_ODD
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill      = make_fill(bg)
            cell.font      = make_font(size=9)
            cell.alignment = make_left()
            cell.border    = make_border()
        ws.cell(row=row, column=1).value = sw.get("name","—")
        ws.cell(row=row, column=2).value = sw.get("version","—")
        ws.cell(row=row, column=3).value = sw.get("publisher","—")
        ws.cell(row=row, column=4).value = sw.get("install_date","—")
        ws.cell(row=row, column=5).value = sw.get("size","—")
        ws.row_dimensions[row].height = 16
        row += 1

    ws.auto_filter.ref = f"A2:E{row-1}"


# ══════════════════════════════════════════════════════
#  LICENÇAS
# ══════════════════════════════════════════════════════

def build_licencas(wb, data):
    ws = wb.create_sheet("Licenças")
    ws.sheet_view.showGridLines = False
    for r in range(1, 50):
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = make_fill(COLOR_BG_DARK)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45

    licenses  = data.get("licenses", {})
    logo_path = get_logo_path()
    add_logo(ws, logo_path, row=1, col=4)

    row = 1
    row = add_section_title(ws, row, "🪟  LICENÇA WINDOWS", 2)
    style_header_row(ws, row, 2)
    ws.cell(row=row, column=1).value = "Campo"
    ws.cell(row=row, column=2).value = "Valor"
    row += 1
    for i, (label, value) in enumerate([
        ("Edição",          licenses.get("windows_edition","—")),
        ("Build",           licenses.get("windows_build","—")),
        ("Status",          licenses.get("windows_status","—")),
        ("Chave (parcial)", licenses.get("windows_key_partial","—")),
        ("Tipo de Licença", licenses.get("windows_license_type","—")),
    ]):
        style_data_row(ws, row, 2, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = make_font(bold=True,
                                                     color=COLOR_TEAL)
        ws.cell(row=row, column=2).value = value
        if label == "Status":
            color = (COLOR_GREEN if value == "Ativado"
                     else COLOR_RED)
            ws.cell(row=row, column=2).font = make_font(
                bold=True, color=color)
        row += 1

    row += 1
    row = add_section_title(ws, row, "📊  LICENÇA OFFICE", 2)
    style_header_row(ws, row, 2)
    ws.cell(row=row, column=1).value = "Campo"
    ws.cell(row=row, column=2).value = "Valor"
    row += 1
    for i, (label, value) in enumerate([
        ("Versão",          licenses.get("office_version","—")),
        ("Produto",         licenses.get("office_product","—")),
        ("Status",          licenses.get("office_status","—")),
        ("Chave (parcial)", licenses.get("office_key_partial","—")),
        ("Conta Vinculada", licenses.get("office_account","—")),
        ("Apps Instalados", ", ".join(
            licenses.get("office_apps", []))),
    ]):
        style_data_row(ws, row, 2, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = make_font(bold=True,
                                                     color=COLOR_TEAL)
        ws.cell(row=row, column=2).value = value
        if label == "Status":
            color = (COLOR_GREEN if value == "Ativado"
                     else COLOR_YELLOW)
            ws.cell(row=row, column=2).font = make_font(
                bold=True, color=color)
        row += 1


# ══════════════════════════════════════════════════════
#  IMPRESSORAS
# ══════════════════════════════════════════════════════

def build_impressoras(wb, data):
    ws = wb.create_sheet("Impressoras")
    ws.sheet_view.showGridLines = False
    for r in range(1, 40):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = make_fill(COLOR_BG_DARK)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    printers  = data.get("printers", {})
    logo_path = get_logo_path()
    add_logo(ws, logo_path, row=1, col=6)

    row = 1
    row = add_section_title(
        ws, row,
        f"🖨️  IMPRESSORAS  —  "
        f"{printers.get('printers_count',0)} encontradas  |  "
        f"Padrão: {printers.get('default_printer','—')}",
        6
    )
    style_header_row(ws, row, 6)
    for col, h in enumerate(
            ["Nome","Padrão","IP","Driver","Status","Compartilhada"],
            1):
        ws.cell(row=row, column=col).value = h
    row += 1

    for i, p in enumerate(printers.get("printers", [])):
        style_data_row(ws, row, 6, even=(i % 2 == 0))
        ws.cell(row=row, column=1).value = p.get("name","—")
        ws.cell(row=row, column=2).value = p.get("default","—")
        ws.cell(row=row, column=3).value = p.get("ip","—")
        ws.cell(row=row, column=4).value = p.get("driver","—")
        status = p.get("status","—")
        ws.cell(row=row, column=5).value = status
        ws.cell(row=row, column=5).font = make_font(
            bold=True,
            color=COLOR_GREEN if status == "Online" else COLOR_RED)
        ws.cell(row=row, column=6).value = p.get("shared","—")
        row += 1


# ══════════════════════════════════════════════════════
#  MAPEAMENTOS
# ══════════════════════════════════════════════════════

def build_mapeamentos(wb, data):
    ws = wb.create_sheet("Mapeamentos")
    ws.sheet_view.showGridLines = False
    for r in range(1, 40):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = make_fill(COLOR_BG_DARK)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    mappings  = data.get("mappings", {})
    logo_path = get_logo_path()
    add_logo(ws, logo_path, row=1, col=6)

    row = 1
    row = add_section_title(
        ws, row, "📁  DRIVES DE REDE MAPEADOS", 6)
    style_header_row(ws, row, 6)
    for col, h in enumerate(
            ["Drive","Caminho UNC","Servidor",
             "Tamanho","Livre","Status"], 1):
        ws.cell(row=row, column=col).value = h
    row += 1

    mapping_list = mappings.get("mappings", [])
    if mapping_list:
        for i, m in enumerate(mapping_list):
            style_data_row(ws, row, 6, even=(i % 2 == 0))
            ws.cell(row=row, column=1).value = m.get("drive","—")
            ws.cell(row=row, column=2).value = m.get("unc_path","—")
            ws.cell(row=row, column=3).value = m.get("server","—")
            ws.cell(row=row, column=4).value = m.get("size","—")
            ws.cell(row=row, column=5).value = m.get("free","—")
            ws.cell(row=row, column=6).value = m.get("status","—")
            row += 1
    else:
        style_data_row(ws, row, 6)
        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row=row, column=1).value = \
            "Nenhum drive mapeado encontrado"
        ws.cell(row=row, column=1).alignment = make_center()
        ws.cell(row=row, column=1).font = make_font(
            color=COLOR_GRAY, size=9)
        row += 1

    row += 1
    row = add_section_title(
        ws, row, "🔗  COMPARTILHAMENTOS LOCAIS", 3)
    style_header_row(ws, row, 3)
    for col, h in enumerate(["Nome","Caminho","Descrição"], 1):
        ws.cell(row=row, column=col).value = h
    row += 1

    shares = mappings.get("local_shares", [])
    if shares:
        for i, s in enumerate(shares):
            style_data_row(ws, row, 3, even=(i % 2 == 0))
            ws.cell(row=row, column=1).value = s.get("name","—")
            ws.cell(row=row, column=2).value = s.get("path","—")
            ws.cell(row=row, column=3).value = s.get(
                "description","—")
            row += 1
    else:
        style_data_row(ws, row, 3)
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row=row, column=1).value = \
            "Nenhum compartilhamento local encontrado"
        ws.cell(row=row, column=1).alignment = make_center()
        ws.cell(row=row, column=1).font = make_font(
            color=COLOR_GRAY, size=9)


# ══════════════════════════════════════════════════════
#  FUNÇÃO PRINCIPAL
# ══════════════════════════════════════════════════════

def generate_xls(data, output_dir=None, filepath=None):
    if filepath is None:
        if output_dir is None:
            base = getattr(sys, '_MEIPASS',
                           os.path.dirname(os.path.abspath(__file__)))
            output_dir = os.path.normpath(os.path.join(base, ".."))
        os.makedirs(output_dir, exist_ok=True)
        machine  = data.get("system",{}).get("machine_name","PC")
        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(
            output_dir, f"inventario_{machine}_{date_str}.xlsx")

    wb = Workbook()
    build_capa(wb, data)
    build_sistema(wb, data)
    build_hardware(wb, data)
    build_rede(wb, data)
    build_software(wb, data)
    build_licencas(wb, data)
    build_impressoras(wb, data)
    build_mapeamentos(wb, data)
    wb.save(filepath)
    return filepath