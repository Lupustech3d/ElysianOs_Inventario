import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from datetime import datetime

BASE_DIR = getattr(sys, '_MEIPASS',
                   os.path.dirname(os.path.abspath(__file__)))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)


# ══════════════════════════════════════════════════════
#  LEITOR DE XLSX
# ══════════════════════════════════════════════════════

def read_xlsx_data(filepath):
    """Lê um arquivo xlsx gerado pelo ElysianOS e extrai os dados"""
    from openpyxl import load_workbook
    result = {
        "filename": os.path.basename(filepath),
        "filepath": filepath,
        "machine":       "—",
        "collaborator":  "—",
        "windows_user":  "—",
        "os":            "—",
        "ip":            "—",
        "serial":        "—",
        "manufacturer":  "—",
        "model":         "—",
        "cpu":           "—",
        "ram":           "—",
        "disk":          "—",
        "windows_status":"—",
        "windows_key":   "—",
        "office_version":"—",
        "office_product":"—",
        "office_status": "—",
        "office_key":    "—",
        "office_apps":   "—",
        "printers":      [],
        "mappings":      [],
        "software":      [],
        "adapters":      [],
    }

    try:
        wb = load_workbook(filepath, data_only=True)

        # ── Aba Sistema ──────────────────────────────
        if "Sistema" in wb.sheetnames:
            ws = wb["Sistema"]
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                label = str(row[0]).strip()
                value = str(row[1]).strip() if row[1] else "—"
                mapping = {
                    "Nome da Máquina":     "machine",
                    "Colaborador":         "collaborator",
                    "Usuário Windows":     "windows_user",
                    "Sistema Operacional": "os",
                    "Fabricante":          "manufacturer",
                    "Modelo":              "model",
                    "Número de Série":     "serial",
                }
                if label in mapping:
                    result[mapping[label]] = value

        # ── Aba Hardware ─────────────────────────────
        if "Hardware" in wb.sheetnames:
            ws = wb["Hardware"]
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                label = str(row[0]).strip()
                value = str(row[1]).strip() if row[1] else "—"
                if label == "Modelo":
                    result["cpu"] = value
                elif label == "Total" and result["ram"] == "—":
                    result["ram"] = value

        # ── Aba Rede ─────────────────────────────────
        if "Rede" in wb.sheetnames:
            ws = wb["Rede"]
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                label = str(row[0]).strip()
                value = str(row[1]).strip() if row[1] else "—"
                if label == "IP Principal":
                    result["ip"] = value

        # ── Aba Licenças ─────────────────────────────
        if "Licenças" in wb.sheetnames:
            ws = wb["Licenças"]
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                label = str(row[0]).strip()
                value = str(row[1]).strip() if row[1] else "—"
                lmap = {
                    "Status":          "windows_status",
                    "Chave (parcial)": "windows_key",
                    "Versão":          "office_version",
                    "Produto":         "office_product",
                    "Apps Instalados": "office_apps",
                }
                # Distingue Windows x Office pelo contexto
                if label == "Status" and result["windows_status"] == "—":
                    result["windows_status"] = value
                elif label == "Chave (parcial)" and result["windows_key"] == "—":
                    result["windows_key"] = value
                elif label == "Status" and result["office_status"] == "—":
                    result["office_status"] = value
                elif label == "Chave (parcial)" and result["office_key"] == "—":
                    result["office_key"] = value
                elif label == "Versão":
                    result["office_version"] = value
                elif label == "Produto":
                    result["office_product"] = value
                elif label == "Apps Instalados":
                    result["office_apps"] = value

        # ── Aba Impressoras ──────────────────────────
        if "Impressoras" in wb.sheetnames:
            ws = wb["Impressoras"]
            printers = []
            header_found = False
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                if str(row[0]).strip() == "Nome":
                    header_found = True
                    continue
                if header_found and row[0]:
                    printers.append({
                        "name":   str(row[0] or "—"),
                        "status": str(row[4] or "—") if len(row) > 4 else "—",
                        "ip":     str(row[2] or "—") if len(row) > 2 else "—",
                    })
            result["printers"] = printers

        # ── Aba Mapeamentos ──────────────────────────
        if "Mapeamentos" in wb.sheetnames:
            ws = wb["Mapeamentos"]
            mappings = []
            header_found = False
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                if str(row[0]).strip() == "Drive":
                    header_found = True
                    continue
                if header_found and row[0]:
                    mappings.append({
                        "drive": str(row[0] or "—"),
                        "path":  str(row[1] or "—") if len(row) > 1 else "—",
                    })
            result["mappings"] = mappings

        # ── Aba Software ─────────────────────────────
        if "Software" in wb.sheetnames:
            ws = wb["Software"]
            software = []
            header_found = False
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                if str(row[0]).strip() == "Nome":
                    header_found = True
                    continue
                if header_found and row[0]:
                    software.append({
                        "name":      str(row[0] or "—"),
                        "version":   str(row[1] or "—") if len(row) > 1 else "—",
                        "publisher": str(row[2] or "—") if len(row) > 2 else "—",
                    })
            result["software"] = software

        wb.close()

    except Exception as e:
        result["error"] = str(e)

    return result


# ══════════════════════════════════════════════════════
#  GERADOR XLS CONSOLIDADO
# ══════════════════════════════════════════════════════

def generate_consolidated_xls(all_data, filepath):
    """Gera XLS consolidado com todas as máquinas"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    C_DARK  = "1A1A2E"
    C_MID   = "16213E"
    C_LIGHT = "0F3460"
    C_TEAL  = "00BCD4"
    C_WHITE = "FFFFFF"
    C_GRAY  = "7F8C8D"
    C_GREEN = "00BC64"
    C_RED   = "E74C3C"
    C_YELLOW= "F39C12"
    C_EVEN  = "141D35"
    C_ODD   = "16213E"

    def fill(c): return PatternFill("solid", fgColor=c)
    def font(bold=False, color=C_WHITE, size=9):
        return Font(bold=bold, color=color, size=size, name="Segoe UI")
    def border():
        s = Side(style="thin", color="1E3A5F")
        return Border(left=s, right=s, top=s, bottom=s)
    def center():
        return Alignment(horizontal="center", vertical="center",
                         wrap_text=True)
    def left():
        return Alignment(horizontal="left", vertical="center",
                         wrap_text=False)

    def hdr_row(ws, row, cols):
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill(C_LIGHT)
            cell.font = font(bold=True, color=C_TEAL)
            cell.alignment = center()
            cell.border = border()
        ws.row_dimensions[row].height = 20

    def data_row(ws, row, cols, even=True):
        bg = C_EVEN if even else C_ODD
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill(bg)
            cell.font = font(size=9)
            cell.alignment = left()
            cell.border = border()
        ws.row_dimensions[row].height = 16

    def section(ws, row, title, cols):
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=cols)
        c = ws.cell(row=row, column=1)
        c.value = title
        c.fill = fill(C_DARK)
        c.font = font(bold=True, color=C_TEAL, size=10)
        c.alignment = left()
        ws.row_dimensions[row].height = 22
        return row + 1

    def add_logo_ws(ws, col=1, row=1):
        logo_path = os.path.join(BASE_DIR, "assets", "logo.png")
        try:
            img = XLImage(logo_path)
            img.width = 45; img.height = 45
            ws.add_image(img, f"{get_column_letter(col)}{row}")
        except:
            pass

    def dark_bg(ws, rows, cols):
        for r in range(1, rows):
            for c in range(1, cols):
                ws.cell(row=r, column=c).fill = fill(C_DARK)

    wb = Workbook()

    # ── ABA RESUMO GERAL ─────────────────────────────
    ws = wb.active
    ws.title = "Resumo Geral"
    ws.sheet_view.showGridLines = False
    dark_bg(ws, 200, 15)
    add_logo_ws(ws, col=12, row=1)

    col_w = [20,22,18,16,22,18,14,14,16,16,14]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Título
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "ElysianOS — Relatório Consolidado"
    c.font = Font(bold=True, color=C_TEAL, size=16, name="Segoe UI")
    c.fill = fill(C_DARK)
    c.alignment = center()
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = (f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
               f"Total de máquinas: {len(all_data)}")
    c.font = Font(color=C_GRAY, size=9, name="Segoe UI")
    c.fill = fill(C_DARK)
    c.alignment = center()
    ws.row_dimensions[2].height = 18

    row = 4
    row = section(ws, row, "📋  RESUMO DE TODAS AS MÁQUINAS", 11)
    hdr_row(ws, row, 11)
    headers = ["Máquina","Colaborador","IP","SO","CPU","RAM",
               "Win Status","Office","Office Status","Impressoras","Mapeamentos"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i).value = h
    row += 1

    for idx, d in enumerate(all_data):
        data_row(ws, row, 11, even=(idx % 2 == 0))
        ws.cell(row=row, column=1).value  = d.get("machine","—")
        ws.cell(row=row, column=2).value  = d.get("collaborator","—")
        ws.cell(row=row, column=3).value  = d.get("ip","—")
        ws.cell(row=row, column=4).value  = d.get("os","—")
        ws.cell(row=row, column=5).value  = d.get("cpu","—")
        ws.cell(row=row, column=6).value  = d.get("ram","—")
        win_s = d.get("windows_status","—")
        ws.cell(row=row, column=7).value  = win_s
        ws.cell(row=row, column=7).font   = font(
            bold=True,
            color=C_GREEN if win_s == "Ativado" else C_RED)
        ws.cell(row=row, column=8).value  = d.get("office_version","—")
        off_s = d.get("office_status","—")
        ws.cell(row=row, column=9).value  = off_s
        ws.cell(row=row, column=9).font   = font(
            bold=True,
            color=C_GREEN if off_s == "Ativado" else C_YELLOW)
        ws.cell(row=row, column=10).value = str(
            len(d.get("printers",[])))
        ws.cell(row=row, column=11).value = str(
            len(d.get("mappings",[])))
        row += 1

    # ── ABA LICENÇAS ─────────────────────────────────
    ws2 = wb.create_sheet("Licenças")
    ws2.sheet_view.showGridLines = False
    dark_bg(ws2, 200, 10)
    add_logo_ws(ws2, col=9, row=1)

    for i, w in enumerate([22,20,16,32,32,20,32,20], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = section(ws2, row, "🔑  LICENÇAS — TODAS AS MÁQUINAS", 8)
    hdr_row(ws2, row, 8)
    for i, h in enumerate(
            ["Máquina","Colaborador","Win Status","Win Chave",
             "Office Produto","Office Status","Office Chave",
             "Apps Office"], 1):
        ws2.cell(row=row, column=i).value = h
    row += 1

    for idx, d in enumerate(all_data):
        data_row(ws2, row, 8, even=(idx % 2 == 0))
        ws2.cell(row=row, column=1).value = d.get("machine","—")
        ws2.cell(row=row, column=2).value = d.get("collaborator","—")
        win_s = d.get("windows_status","—")
        ws2.cell(row=row, column=3).value = win_s
        ws2.cell(row=row, column=3).font  = font(
            bold=True,
            color=C_GREEN if win_s == "Ativado" else C_RED)
        ws2.cell(row=row, column=4).value = d.get("windows_key","—")
        ws2.cell(row=row, column=5).value = d.get("office_product","—")
        off_s = d.get("office_status","—")
        ws2.cell(row=row, column=6).value = off_s
        ws2.cell(row=row, column=6).font  = font(
            bold=True,
            color=C_GREEN if off_s == "Ativado" else C_YELLOW)
        ws2.cell(row=row, column=7).value = d.get("office_key","—")
        ws2.cell(row=row, column=8).value = d.get("office_apps","—")
        row += 1

    # ── ABA HARDWARE ─────────────────────────────────
    ws3 = wb.create_sheet("Hardware")
    ws3.sheet_view.showGridLines = False
    dark_bg(ws3, 200, 10)
    add_logo_ws(ws3, col=8, row=1)

    for i, w in enumerate([22,20,14,40,14,14,20,20], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = section(ws3, row, "🔧  HARDWARE — TODAS AS MÁQUINAS", 8)
    hdr_row(ws3, row, 8)
    for i, h in enumerate(
            ["Máquina","Colaborador","IP","CPU",
             "RAM","Disco","Fabricante","Modelo"], 1):
        ws3.cell(row=row, column=i).value = h
    row += 1

    for idx, d in enumerate(all_data):
        data_row(ws3, row, 8, even=(idx % 2 == 0))
        ws3.cell(row=row, column=1).value = d.get("machine","—")
        ws3.cell(row=row, column=2).value = d.get("collaborator","—")
        ws3.cell(row=row, column=3).value = d.get("ip","—")
        ws3.cell(row=row, column=4).value = d.get("cpu","—")
        ws3.cell(row=row, column=5).value = d.get("ram","—")
        ws3.cell(row=row, column=6).value = d.get("disk","—")
        ws3.cell(row=row, column=7).value = d.get("manufacturer","—")
        ws3.cell(row=row, column=8).value = d.get("model","—")
        row += 1

    # ── ABA SOFTWARE ─────────────────────────────────
    ws4 = wb.create_sheet("Softwares")
    ws4.sheet_view.showGridLines = False
    dark_bg(ws4, 5, 6)
    add_logo_ws(ws4, col=5, row=1)

    for i, w in enumerate([22,45,18,30], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = section(ws4, row, "📦  SOFTWARES — TODAS AS MÁQUINAS", 4)
    hdr_row(ws4, row, 4)
    for i, h in enumerate(
            ["Máquina","Software","Versão","Fabricante"], 1):
        ws4.cell(row=row, column=i).value = h
    row += 1

    for idx_m, d in enumerate(all_data):
        for idx_s, sw in enumerate(d.get("software",[])):
            even = (row % 2 == 0)
            data_row(ws4, row, 4, even=even)
            ws4.cell(row=row, column=1).value = d.get("machine","—")
            ws4.cell(row=row, column=2).value = sw.get("name","—")
            ws4.cell(row=row, column=3).value = sw.get("version","—")
            ws4.cell(row=row, column=4).value = sw.get("publisher","—")
            row += 1

    ws4.auto_filter.ref = f"A2:D{row-1}"

    # ── ABA IMPRESSORAS ──────────────────────────────
    ws5 = wb.create_sheet("Impressoras")
    ws5.sheet_view.showGridLines = False
    dark_bg(ws5, 5, 6)
    add_logo_ws(ws5, col=5, row=1)

    for i, w in enumerate([22,20,38,14,18], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = section(ws5, row, "🖨️  IMPRESSORAS — TODAS AS MÁQUINAS", 5)
    hdr_row(ws5, row, 5)
    for i, h in enumerate(
            ["Máquina","Colaborador","Impressora","IP","Status"], 1):
        ws5.cell(row=row, column=i).value = h
    row += 1

    for d in all_data:
        for p in d.get("printers",[]):
            even = (row % 2 == 0)
            data_row(ws5, row, 5, even=even)
            ws5.cell(row=row, column=1).value = d.get("machine","—")
            ws5.cell(row=row, column=2).value = d.get("collaborator","—")
            ws5.cell(row=row, column=3).value = p.get("name","—")
            ws5.cell(row=row, column=4).value = p.get("ip","—")
            status = p.get("status","—")
            ws5.cell(row=row, column=5).value = status
            ws5.cell(row=row, column=5).font  = font(
                bold=True,
                color=C_GREEN if status == "Online" else C_RED)
            row += 1

    # ── ABA MAPEAMENTOS ──────────────────────────────
    ws6 = wb.create_sheet("Mapeamentos")
    ws6.sheet_view.showGridLines = False
    dark_bg(ws6, 5, 5)
    add_logo_ws(ws6, col=4, row=1)

    for i, w in enumerate([22,20,12,38], 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = section(ws6, row, "📁  MAPEAMENTOS — TODAS AS MÁQUINAS", 4)
    hdr_row(ws6, row, 4)
    for i, h in enumerate(
            ["Máquina","Colaborador","Drive","Caminho UNC"], 1):
        ws6.cell(row=row, column=i).value = h
    row += 1

    for d in all_data:
        for m in d.get("mappings",[]):
            even = (row % 2 == 0)
            data_row(ws6, row, 4, even=even)
            ws6.cell(row=row, column=1).value = d.get("machine","—")
            ws6.cell(row=row, column=2).value = d.get("collaborator","—")
            ws6.cell(row=row, column=3).value = m.get("drive","—")
            ws6.cell(row=row, column=4).value = m.get("path","—")
            row += 1

    wb.save(filepath)
    return filepath


# ══════════════════════════════════════════════════════
#  RELATÓRIO RÁPIDO — Máquina + Colaborador + Licença
# ══════════════════════════════════════════════════════

def generate_quick_report_xls(all_data, filepath):
    """Gera lista simplificada: Máquina | Colaborador | Licença Office"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_DARK  = "1A1A2E"
    C_LIGHT = "0F3460"
    C_TEAL  = "00BCD4"
    C_WHITE = "FFFFFF"
    C_GRAY  = "7F8C8D"
    C_GREEN = "00BC64"
    C_YELLOW= "F39C12"
    C_RED   = "E74C3C"
    C_EVEN  = "141D35"
    C_ODD   = "16213E"

    def fill(c): return PatternFill("solid", fgColor=c)
    def font(bold=False, color=C_WHITE, size=9):
        return Font(bold=bold, color=color, size=size, name="Segoe UI")
    def border():
        s = Side(style="thin", color="1E3A5F")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Rápido"
    ws.sheet_view.showGridLines = False

    for r in range(1, len(all_data) + 10):
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = fill(C_DARK)

    # Larguras
    for i, w in enumerate([22,22,20,32,20,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Título
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "ElysianOS — Relatório Rápido de Licenças"
    c.font = Font(bold=True, color=C_TEAL, size=14, name="Segoe UI")
    c.fill = fill(C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = (f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  "
               f"|  Total: {len(all_data)} máquinas")
    c.font = Font(color=C_GRAY, size=8, name="Segoe UI")
    c.fill = fill(C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Cabeçalho
    row = 4
    headers = ["Máquina","Colaborador","Usuário Windows",
               "Office Produto","Office Status","Chave Office"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill(C_LIGHT)
        cell.font = font(bold=True, color=C_TEAL)
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center")
        cell.border = border()
    ws.row_dimensions[row].height = 20
    row += 1

    for idx, d in enumerate(all_data):
        bg = C_EVEN if idx % 2 == 0 else C_ODD
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill(bg)
            cell.font = font(size=9)
            cell.alignment = Alignment(horizontal="left",
                                       vertical="center")
            cell.border = border()
        ws.cell(row=row, column=1).value = d.get("machine","—")
        ws.cell(row=row, column=2).value = d.get("collaborator","—")
        ws.cell(row=row, column=3).value = d.get("windows_user","—")
        ws.cell(row=row, column=4).value = d.get("office_product","—")
        off_s = d.get("office_status","—")
        ws.cell(row=row, column=5).value = off_s
        ws.cell(row=row, column=5).font  = font(
            bold=True,
            color=C_GREEN if off_s == "Ativado" else C_YELLOW)
        ws.cell(row=row, column=6).value = d.get("office_key","—")
        ws.row_dimensions[row].height = 16
        row += 1

    ws.auto_filter.ref = f"A4:F{row-1}"
    wb.save(filepath)
    return filepath


# ══════════════════════════════════════════════════════
#  PDF CONSOLIDADO
# ══════════════════════════════════════════════════════

def generate_consolidated_pdf(all_data, filepath):
    """Gera PDF consolidado com todas as máquinas"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table,
        TableStyle, HRFlowable, PageBreak
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    from reportlab.platypus import Image as RLImage

    C_DARK  = colors.HexColor("#1A1A2E")
    C_MID   = colors.HexColor("#16213E")
    C_LIGHT = colors.HexColor("#0F3460")
    C_TEAL  = colors.HexColor("#00BCD4")
    C_WHITE = colors.white
    C_GRAY  = colors.HexColor("#7F8C8D")
    C_GREEN = colors.HexColor("#00BC64")
    C_RED   = colors.HexColor("#E74C3C")
    C_YELLOW= colors.HexColor("#F39C12")
    C_EVEN  = colors.HexColor("#141D35")
    C_ODD   = colors.HexColor("#16213E")

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="ElyTitle", fontName="Helvetica-Bold", fontSize=20,
        textColor=C_TEAL, alignment=TA_CENTER, spaceAfter=4))
    styles.add(ParagraphStyle(
        name="ElySubtitle", fontName="Helvetica", fontSize=10,
        textColor=C_GRAY, alignment=TA_CENTER, spaceAfter=2))
    styles.add(ParagraphStyle(
        name="ElySection", fontName="Helvetica-Bold", fontSize=11,
        textColor=C_TEAL, alignment=TA_LEFT,
        spaceBefore=10, spaceAfter=5))
    styles.add(ParagraphStyle(
        name="ElyNormal", fontName="Helvetica", fontSize=8,
        textColor=C_WHITE, alignment=TA_LEFT))
    styles.add(ParagraphStyle(
        name="ElySmall", fontName="Helvetica", fontSize=7,
        textColor=C_GRAY, alignment=TA_CENTER))

    pw = A4[0] - 30*mm

    def header_footer(canvas, doc):
        canvas.saveState()
        w, h = A4
        canvas.setFillColor(C_DARK)
        canvas.rect(0, h - 40*mm, w, 40*mm, fill=1, stroke=0)
        logo_path = os.path.join(BASE_DIR, "assets", "logo.png")
        try:
            canvas.drawImage(logo_path, 15*mm, h-34*mm,
                             width=18*mm, height=18*mm,
                             preserveAspectRatio=True, mask="auto")
        except:
            pass
        canvas.setFillColor(C_TEAL)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawString(38*mm, h-18*mm, "ElysianOS")
        canvas.setFillColor(C_GRAY)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(38*mm, h-24*mm,
                          "Sistema de Inventário Integrado — Consolidado")
        canvas.setFillColor(C_WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(w-15*mm, h-18*mm,
                               f"{len(all_data)} máquinas")
        canvas.drawRightString(w-15*mm, h-24*mm,
            datetime.now().strftime("%d/%m/%Y %H:%M"))
        canvas.setStrokeColor(C_TEAL)
        canvas.setLineWidth(1.5)
        canvas.line(0, h-40*mm, w, h-40*mm)
        canvas.setFillColor(C_MID)
        canvas.rect(0, 0, w, 12*mm, fill=1, stroke=0)
        canvas.setStrokeColor(C_TEAL)
        canvas.setLineWidth(0.5)
        canvas.line(0, 12*mm, w, 12*mm)
        canvas.setFillColor(C_GRAY)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(15*mm, 4*mm,
            f"ElysianOS {datetime.now().year} | Desenvolvido por Luis Mesquitela")
        canvas.drawRightString(w-15*mm, 4*mm, f"Página {doc.page}")
        canvas.restoreState()

    def make_table(headers, rows, col_widths):
        clean = [[str(c) if c else "—" for c in r] for r in rows]
        data  = [headers] + clean
        style = TableStyle([
            ("BACKGROUND",    (0,0),(-1,0),  C_LIGHT),
            ("TEXTCOLOR",     (0,0),(-1,0),  C_TEAL),
            ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,0),  8),
            ("ALIGN",         (0,0),(-1,0),  "CENTER"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [C_EVEN, C_ODD]),
            ("TEXTCOLOR",     (0,1),(-1,-1), C_WHITE),
            ("FONTNAME",      (0,1),(-1,-1), "Helvetica"),
            ("FONTSIZE",      (0,1),(-1,-1), 7),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("GRID",          (0,0),(-1,-1), 0.3,
             colors.HexColor("#1E3A5F")),
            ("TOPPADDING",    (0,0),(-1,-1), 3),
            ("BOTTOMPADDING", (0,0),(-1,-1), 3),
            ("LEFTPADDING",   (0,0),(-1,-1), 4),
        ])
        return Table(data, colWidths=col_widths,
                     style=style, repeatRows=1, hAlign="LEFT")

    story = []

    # Capa
    story.append(Spacer(1, 15*mm))
    logo_path = os.path.join(BASE_DIR, "assets", "logo.png")
    try:
        logo = RLImage(logo_path, width=45*mm, height=45*mm)
        logo.hAlign = "CENTER"
        story.append(logo)
    except:
        pass
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph("ElysianOS", styles["ElyTitle"]))
    story.append(Paragraph("Relatório Consolidado de Inventário",
                            styles["ElySubtitle"]))
    story.append(Spacer(1, 5*mm))
    story.append(HRFlowable(width="100%", thickness=1.5,
                             color=C_TEAL, spaceAfter=5*mm))
    story.append(Paragraph(
        f"Total de máquinas: {len(all_data)}<br/>"
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles["ElySubtitle"]))
    story.append(Spacer(1, 5*mm))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=C_TEAL, spaceAfter=3*mm))
    story.append(Paragraph("Desenvolvido por Luis Mesquitela",
                            styles["ElySmall"]))
    story.append(PageBreak())

    # Resumo geral
    story.append(Paragraph("▌ RESUMO GERAL", styles["ElySection"]))
    story.append(make_table(
        ["Máquina","Colaborador","IP","SO","Win","Office","Status Office"],
        [[d.get("machine","—"), d.get("collaborator","—"),
          d.get("ip","—"),      d.get("os","—"),
          d.get("windows_status","—"),
          d.get("office_version","—"),
          d.get("office_status","—")]
         for d in all_data],
        col_widths=[pw*0.16, pw*0.16, pw*0.12, pw*0.18,
                    pw*0.12, pw*0.14, pw*0.12]
    ))
    story.append(Spacer(1, 5*mm))

    # Licenças
    story.append(Paragraph("▌ LICENÇAS", styles["ElySection"]))
    story.append(make_table(
        ["Máquina","Colaborador","Win Status","Office Produto",
         "Office Status","Chave Office"],
        [[d.get("machine","—"),       d.get("collaborator","—"),
          d.get("windows_status","—"),d.get("office_product","—"),
          d.get("office_status","—"), d.get("office_key","—")]
         for d in all_data],
        col_widths=[pw*0.16, pw*0.16, pw*0.12,
                    pw*0.24, pw*0.14, pw*0.18]
    ))
    story.append(PageBreak())

    # Hardware
    story.append(Paragraph("▌ HARDWARE", styles["ElySection"]))
    story.append(make_table(
        ["Máquina","Colaborador","CPU","RAM","Fabricante","Modelo"],
        [[d.get("machine","—"),      d.get("collaborator","—"),
          d.get("cpu","—"),          d.get("ram","—"),
          d.get("manufacturer","—"), d.get("model","—")]
         for d in all_data],
        col_widths=[pw*0.16, pw*0.16, pw*0.26,
                    pw*0.10, pw*0.16, pw*0.16]
    ))
    story.append(Spacer(1, 5*mm))

    # Impressoras
    all_printers = []
    for d in all_data:
        for p in d.get("printers",[]):
            all_printers.append([
                d.get("machine","—"),
                d.get("collaborator","—"),
                p.get("name","—"),
                p.get("ip","—"),
                p.get("status","—")
            ])
    if all_printers:
        story.append(Paragraph("▌ IMPRESSORAS", styles["ElySection"]))
        story.append(make_table(
            ["Máquina","Colaborador","Impressora","IP","Status"],
            all_printers,
            col_widths=[pw*0.18, pw*0.16, pw*0.34,
                        pw*0.16, pw*0.16]
        ))
        story.append(Spacer(1, 5*mm))

    # Mapeamentos
    all_maps = []
    for d in all_data:
        for m in d.get("mappings",[]):
            all_maps.append([
                d.get("machine","—"),
                d.get("collaborator","—"),
                m.get("drive","—"),
                m.get("path","—")
            ])
    if all_maps:
        story.append(Paragraph("▌ MAPEAMENTOS", styles["ElySection"]))
        story.append(make_table(
            ["Máquina","Colaborador","Drive","Caminho UNC"],
            all_maps,
            col_widths=[pw*0.20, pw*0.20, pw*0.10, pw*0.50]
        ))

    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        topMargin=45*mm, bottomMargin=18*mm,
        leftMargin=15*mm, rightMargin=15*mm,
        title="ElysianOS - Relatório Consolidado",
        author="Luis Mesquitela"
    )
    doc.build(story,
              onFirstPage=header_footer,
              onLaterPages=header_footer)
    return filepath


# ══════════════════════════════════════════════════════
#  INTERFACE GRÁFICA
# ══════════════════════════════════════════════════════

class ConsolidatorUI:
    def __init__(self):
        self.files = []
        self.root  = tk.Tk()
        self.root.title("ElysianOS — Consolidador")
        self.root.resizable(False, False)
        self.root.configure(bg="#1a1a2e")
        self.root.attributes("-topmost", True)

        w, h = 680, 560
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(
            f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        self._build_ui()

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self.root, bg="#16213e", pady=12)
        hdr.pack(fill="x")

        try:
            from PIL import Image, ImageTk
            logo_path = os.path.join(BASE_DIR, "assets", "logo.png")
            img = Image.open(logo_path).resize((40,40), Image.LANCZOS)
            self.logo = ImageTk.PhotoImage(img)
            tk.Label(hdr, image=self.logo,
                     bg="#16213e").pack(side="left", padx=15)
        except:
            tk.Label(hdr, text="⬡", font=("Arial",26),
                     fg="#00BCD4", bg="#16213e").pack(
                         side="left", padx=15)

        title_frame = tk.Frame(hdr, bg="#16213e")
        title_frame.pack(side="left")
        tk.Label(title_frame, text="ElysianOS — Consolidador",
                 font=("Segoe UI",14,"bold"),
                 fg="#00BCD4", bg="#16213e").pack(anchor="w")
        tk.Label(title_frame,
                 text="Unifica múltiplos relatórios em um só",
                 font=("Segoe UI",8),
                 fg="#7f8c8d", bg="#16213e").pack(anchor="w")

        tk.Frame(self.root, bg="#00BCD4", height=1).pack(fill="x")

        # Botões de arquivo
        file_frame = tk.Frame(self.root, bg="#1a1a2e", pady=10)
        file_frame.pack(fill="x", padx=20)

        tk.Button(file_frame,
                  text="➕  Adicionar arquivos XLS",
                  font=("Segoe UI",10,"bold"),
                  bg="#0f3460", fg="#00BCD4",
                  activebackground="#00BCD4",
                  activeforeground="#0f0f1a",
                  relief="flat", padx=15, pady=8,
                  cursor="hand2",
                  command=self._add_files).pack(
                      side="left", padx=(0,10))

        tk.Button(file_frame,
                  text="🗑️  Limpar lista",
                  font=("Segoe UI",10),
                  bg="#1a1a2e", fg="#7f8c8d",
                  activebackground="#16213e",
                  relief="flat", padx=15, pady=8,
                  cursor="hand2",
                  command=self._clear_files).pack(side="left")

        self.count_label = tk.Label(
            file_frame,
            text="0 arquivos selecionados",
            font=("Segoe UI",9),
            fg="#7f8c8d", bg="#1a1a2e")
        self.count_label.pack(side="right")

        # Lista de arquivos
        list_frame = tk.Frame(self.root, bg="#16213e",
                              relief="flat", bd=1)
        list_frame.pack(fill="both", expand=True,
                        padx=20, pady=(0,10))

        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")

        self.file_list = tk.Listbox(
            list_frame,
            font=("Segoe UI",9),
            bg="#16213e", fg="#ecf0f1",
            selectbackground="#0f3460",
            selectforeground="#00BCD4",
            activestyle="none",
            relief="flat",
            yscrollcommand=scrollbar.set
        )
        self.file_list.pack(fill="both", expand=True, padx=5, pady=5)
        scrollbar.config(command=self.file_list.yview)

        # Separador
        tk.Frame(self.root, bg="#1e3a5f", height=1).pack(
            fill="x", padx=20)

        # Status
        self.status_var = tk.StringVar(
            value="Selecione os arquivos XLS para consolidar")
        tk.Label(self.root,
                 textvariable=self.status_var,
                 font=("Segoe UI",9),
                 fg="#7f8c8d", bg="#1a1a2e").pack(pady=8)

        # Progress bar
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("C.Horizontal.TProgressbar",
                        troughcolor="#16213e",
                        background="#00BCD4",
                        thickness=8)
        self.progress = ttk.Progressbar(
            self.root, length=640, mode="determinate",
            style="C.Horizontal.TProgressbar")
        self.progress.pack(padx=20, pady=(0,10))

        # Botões de ação
        btn_frame = tk.Frame(self.root, bg="#1a1a2e", pady=10)
        btn_frame.pack(fill="x", padx=20)

        buttons = [
            ("📊  Consolidar XLS",  "#1d6f42", self._run_xls),
            ("📄  Consolidar PDF",  "#c0392b", self._run_pdf),
            ("📋  Relatório Rápido","#0f3460", self._run_quick),
        ]
        for text, color, cmd in buttons:
            tk.Button(btn_frame, text=text,
                      font=("Segoe UI",10,"bold"),
                      bg=color, fg="#ffffff",
                      activebackground="#ffffff",
                      activeforeground=color,
                      relief="flat", padx=12, pady=10,
                      cursor="hand2",
                      command=cmd).pack(
                          side="left", padx=5, expand=True,
                          fill="x")

        # Rodapé
        tk.Label(self.root,
                 text="Desenvolvido por Luis Mesquitela",
                 font=("Segoe UI",8),
                 fg="#4a4a6a", bg="#1a1a2e").pack(
                     side="bottom", pady=8)

    def _add_files(self):
        files = filedialog.askopenfilenames(
            title="Selecionar relatórios XLS",
            filetypes=[("Excel files","*.xlsx"),
                       ("All files","*.*")]
        )
        for f in files:
            if f not in self.files:
                self.files.append(f)
                self.file_list.insert(
                    "end",
                    f"  📄  {os.path.basename(f)}")
        self.count_label.config(
            text=f"{len(self.files)} arquivo(s) selecionado(s)",
            fg="#00BCD4" if self.files else "#7f8c8d")

    def _clear_files(self):
        self.files.clear()
        self.file_list.delete(0, "end")
        self.count_label.config(
            text="0 arquivos selecionados", fg="#7f8c8d")

    def _validate(self):
        if not self.files:
            messagebox.showwarning(
                "ElysianOS",
                "Selecione pelo menos um arquivo XLS!")
            return False
        return True

    def _set_status(self, msg, color="#7f8c8d"):
        self.status_var.set(msg)
        self.root.update_idletasks()

    def _run_xls(self):
        if not self._validate():
            return
        filepath = filedialog.asksaveasfilename(
            title="Salvar XLS Consolidado",
            defaultextension=".xlsx",
            filetypes=[("Excel files","*.xlsx")],
            initialfile=(f"consolidado_"
                         f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                         f".xlsx")
        )
        if not filepath:
            return
        self._run_task(
            lambda data: generate_consolidated_xls(data, filepath),
            filepath, "XLS")

    def _run_pdf(self):
        if not self._validate():
            return
        filepath = filedialog.asksaveasfilename(
            title="Salvar PDF Consolidado",
            defaultextension=".pdf",
            filetypes=[("PDF files","*.pdf")],
            initialfile=(f"consolidado_"
                         f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                         f".pdf")
        )
        if not filepath:
            return
        self._run_task(
            lambda data: generate_consolidated_pdf(data, filepath),
            filepath, "PDF")

    def _run_quick(self):
        if not self._validate():
            return
        filepath = filedialog.asksaveasfilename(
            title="Salvar Relatório Rápido",
            defaultextension=".xlsx",
            filetypes=[("Excel files","*.xlsx")],
            initialfile=(f"relatorio_rapido_"
                         f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                         f".xlsx")
        )
        if not filepath:
            return
        self._run_task(
            lambda data: generate_quick_report_xls(data, filepath),
            filepath, "Relatório Rápido")

    def _run_task(self, task_fn, filepath, label):
        def run():
            try:
                total = len(self.files)
                all_data = []

                for i, f in enumerate(self.files):
                    self._set_status(
                        f"Lendo {os.path.basename(f)}...",
                        "#00BCD4")
                    self.progress["value"] = (i / total) * 80
                    self.root.update_idletasks()
                    all_data.append(read_xlsx_data(f))

                self._set_status(f"Gerando {label}...", "#00BCD4")
                self.progress["value"] = 90
                self.root.update_idletasks()

                task_fn(all_data)

                self.progress["value"] = 100
                self._set_status(
                    f"✅ {label} gerado com sucesso!", "#00BC64")
                os.startfile(filepath)

            except Exception as e:
                self._set_status(f"❌ Erro: {e}", "#e74c3c")
                messagebox.showerror("ElysianOS — Erro", str(e))
            finally:
                self.root.after(
                    3000,
                    lambda: self.progress.configure(value=0))

        threading.Thread(target=run, daemon=True).start()

    def run(self):
        self.root.mainloop()


# ══════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════

def run_consolidator():
    """Chamado pelo main.py quando clica em Consolidar"""
    app = ConsolidatorUI()
    app.run()


if __name__ == "__main__":
    run_consolidator()